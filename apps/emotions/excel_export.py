# -*- coding: utf-8 -*-
"""
excel_export.py — Exportaciones a Excel profesionales para Conéctate.
Genera archivos .xlsx con diseño institucional, colores, gráficas y formato limpio.
Requiere: openpyxl (incluido en requirements.txt)
"""
import io
from datetime import date
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, LineChart


# ── Paleta institucional ───────────────────────────────────────────────────────
C = {
    'purple':   '7C6EFF', 'purple_l': 'EEE8FF',
    'pink':     'FF6B9D', 'pink_l':   'FFE8F3',
    'dark':     '1E1B4B', 'white':    'FFFFFF',
    'green':    '00B87A', 'green_l':  'E6FBF4',
    'orange':   'FD7E14', 'orange_l': 'FFF3E8',
    'red':      'E0284F', 'red_l':    'FFF0F3',
    'yellow':   'F5B800', 'yellow_l': 'FFF9E6',
    'gray':     '6B7280', 'gray_l':   'F9FAFB',
    'text':     '1F2937', 'muted':    '9CA3AF',
    'border':   'E5E7EB',
}

EMO_LABELS = {
    'feliz':     '😊 Feliz',
    'tranquilo': '😌 Tranquilo',
    'estresado': '😰 Estresado',
    'triste':    '😢 Triste',
    'enojado':   '😠 Enojado',
}
EMO_BG = {
    'feliz': 'FFF9E6', 'tranquilo': 'E6FBF4',
    'estresado': 'FFF3E8', 'triste': 'EEF2FF', 'enojado': 'FFF0F3',
}
EMO_FG = {
    'feliz': 'F5B800', 'tranquilo': '00B87A',
    'estresado': 'FD7E14', 'triste': '3B82F6', 'enojado': 'E0284F',
}


# ── Helpers de estilo ─────────────────────────────────────────────────────────
def _fill(hex6):
    return PatternFill('solid', start_color=hex6, end_color=hex6)


def _font(bold=False, size=9, color=None, name='Arial', italic=False):
    return Font(name=name, bold=bold, size=size,
                color=color or C['text'], italic=italic)


def _align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _border_thin():
    s = Side(style='thin', color=C['border'])
    return Border(left=s, right=s, top=s, bottom=s)


def _border_bottom():
    return Border(bottom=Side(style='hair', color=C['border']))


def _prom_color(p):
    if p >= 3.8: return C['green']
    if p >= 3.0: return C['yellow']
    if p >= 2.0: return C['orange']
    return C['red']


def _tend_label(t):
    if t > 0.3:  return '📈 Mejorando'
    if t < -0.3: return '📉 Bajando'
    return '➡️ Estable'


def _row_bg(stats, idx):
    """Color de fondo de fila según estado del estudiante."""
    if stats['alertas_activas'] > 0:
        return C['red_l']
    if stats['promedio'] >= 4.0 and stats['tendencia'] > 0:
        return C['green_l']
    return C['gray_l'] if idx % 2 == 0 else C['white']


def _set(ws, row, col, value, bold=False, size=9, color=None, bg=None,
         h='center', wrap=False, fmt=None, border=None):
    """Escribe un valor con formato en una celda."""
    from openpyxl.cell.cell import MergedCell
    c = ws.cell(row=row, column=col)
    if isinstance(c, MergedCell):
        return  # No se puede escribir en celdas fusionadas secundarias
    c.value = value
    c.font = _font(bold=bold, size=size, color=color or C['text'])
    c.alignment = _align(h=h, v='center', wrap=wrap)
    if bg:
        c.fill = _fill(bg)
    if fmt:
        c.number_format = fmt
    if border:
        c.border = border
    return c


# ══════════════════════════════════════════════════════════════════════════════
# EXPORTACIÓN PARA PROFESORES — informe completo del curso
# ══════════════════════════════════════════════════════════════════════════════

def generar_excel_profesor(curso, stats_lista, dias=30):
    """
    Genera un libro Excel profesional con el informe del curso.
    Retorna un BytesIO listo para enviar como respuesta HTTP.
    """
    wb = Workbook()

    # ── Hoja 1: Resumen ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Resumen General'
    ws1.sheet_view.showGridLines = False
    ws1.sheet_properties.tabColor = '7C6EFF'

    # Anchos de columna
    for col, w in zip('ABCDEFGHI', [3, 26, 13, 13, 13, 15, 13, 13, 4]):
        ws1.column_dimensions[col].width = w

    # ─ Banner ─
    ws1.row_dimensions[1].height = 12
    ws1.row_dimensions[2].height = 20
    ws1.row_dimensions[3].height = 32
    ws1.row_dimensions[4].height = 18
    ws1.row_dimensions[5].height = 16
    ws1.row_dimensions[6].height = 12

    for merge in ['B2:H2', 'B3:H3', 'B4:H4', 'B5:H5']:
        ws1.merge_cells(merge)

    for row in range(2, 6):
        for col in range(2, 9):
            ws1.cell(row=row, column=col).fill = _fill(C['dark'])

    prof_name = curso.profesor.get_full_name() if curso.profesor else 'Sin asignar'
    hoy = date.today()

    _set(ws1, 2, 2, 'CONÉCTATE  —  Sistema de Bienestar Emocional',
         size=10, color=C['purple'], h='center', bg=C['dark'])
    _set(ws1, 3, 2, f'Informe Emocional del Curso {curso.nombre}',
         bold=True, size=18, color=C['white'], h='center', bg=C['dark'])
    _set(ws1, 4, 2, f'Profesor/a: {prof_name}   ·   Período: últimos {dias} días   ·   {hoy.strftime("%d/%m/%Y")}',
         size=10, color='ADBBC4', h='center', bg=C['dark'])
    _set(ws1, 5, 2, f'{len(stats_lista)} estudiantes con registros  ·  Código de grupo: {curso.codigo}',
         size=9, color='7B8D9A', h='center', bg=C['dark'])

    # ─ KPI cards (fila 7-10) ─
    todos_prom = [s['promedio'] for s in stats_lista]
    prom_curso = round(sum(todos_prom) / len(todos_prom), 2) if todos_prom else 0
    n_alertas  = sum(1 for s in stats_lista if s['alertas_activas'] > 0)
    n_baja     = sum(1 for s in stats_lista if s['tendencia'] < -0.3)
    n_mejora   = sum(1 for s in stats_lista if s['tendencia'] > 0.3)

    from apps.courses.models import Inscripcion
    total_inscritos = Inscripcion.objects.filter(curso=curso, activa=True).count()
    participacion = round(len(stats_lista) / max(total_inscritos, 1) * 100)

    kpis = [
        ('B', f'{prom_curso}/5',     'Promedio del grupo',     C['purple'], C['purple_l']),
        ('D', f'{n_alertas}',        'Con alertas activas',    C['red'],    C['red_l']),
        ('F', f'{participacion}%',   'Participación',          C['green'],  C['green_l']),
        ('H', f'{n_mejora}',         'Tendencia positiva',     C['yellow'], C['yellow_l']),
    ]

    for r in [7, 8, 9, 10, 11]:
        ws1.row_dimensions[r].height = 22 if r in [8, 9] else 10 if r == 11 else 8

    for col_letter, value, label, fg, bg in kpis:
        end_col = chr(ord(col_letter) + 1)
        for merge_row in [7, 8, 9, 10]:
            ws1.merge_cells(f'{col_letter}{merge_row}:{end_col}{merge_row}')

        for row in [7, 8, 9]:
            for c_off in range(2):
                ws1.cell(row=row, column=ord(col_letter)-64+c_off).fill = _fill(bg)

        _set(ws1, 7, ord(col_letter)-63, label, size=8, color=C['gray'], h='center', bg=bg)
        _set(ws1, 8, ord(col_letter)-63, value, bold=True, size=20, color=fg, h='center', bg=bg)
        # Barra de acento inferior
        for c_off in range(2):
            ws1.cell(row=10, column=ord(col_letter)-64+c_off).fill = _fill(fg)

    ws1.row_dimensions[12].height = 14

    # ─ Tabla de estudiantes ─
    ws1.row_dimensions[13].height = 26
    headers = ['', 'Estudiante', 'Promedio', 'Registros', 'Emoción Dom.', 'Tendencia', 'Días Bajos', 'Alertas']
    for col_i, h in enumerate(headers):
        c = ws1.cell(row=13, column=col_i+1)
        c.value = h
        c.font = _font(bold=True, size=9, color=C['white'])
        c.fill = _fill(C['dark'])
        c.alignment = _align(h='center' if col_i != 1 else 'left', wrap=True)
        if col_i > 0:
            c.border = _border_thin()

    sorted_stats = sorted(stats_lista, key=lambda x: x['promedio'])

    for i, s in enumerate(sorted_stats):
        row = 14 + i
        ws1.row_dimensions[row].height = 22
        bg = _row_bg(s, i)
        pc = _prom_color(s['promedio'])

        # Acento lateral
        ws1.cell(row=row, column=1).fill = _fill(
            C['red'] if s['alertas_activas'] > 0
            else C['green'] if s['promedio'] >= 4.0
            else C['gray_l'])

        # Nombre
        _set(ws1, row, 2, s['nombre'], bold=True, size=9, color=C['text'],
             h='left', bg=bg, border=_border_bottom())

        # Promedio (en color)
        c = ws1.cell(row=row, column=3)
        c.value = s['promedio']
        c.number_format = '0.00'
        c.font = _font(bold=True, size=11, color=pc)
        c.alignment = _align(h='center')
        c.fill = _fill(bg)
        c.border = _border_bottom()

        # Registros
        _set(ws1, row, 4, s['total_registros'], size=9, h='center', bg=bg, border=_border_bottom())

        # Emoción dominante
        _set(ws1, row, 5, EMO_LABELS.get(s['emocion_dominante'], '—'), size=9, h='center', bg=bg, border=_border_bottom())

        # Tendencia
        _set(ws1, row, 6, _tend_label(s['tendencia']), size=9, h='center', bg=bg, border=_border_bottom())

        # Días bajos
        dias_color = C['red'] if s['dias_bajos'] > 3 else C['text']
        _set(ws1, row, 7, s['dias_bajos'], bold=s['dias_bajos'] > 3, size=9,
             color=dias_color, h='center', bg=bg, border=_border_bottom())

        # Alertas
        if s['alertas_activas'] > 0:
            _set(ws1, row, 8, f"⚠ {s['alertas_activas']}", bold=True, size=9,
                 color=C['red'], h='center', bg=C['red_l'], border=_border_bottom())
        else:
            _set(ws1, row, 8, '✓', size=9, color=C['green'], h='center', bg=bg, border=_border_bottom())

    # Totals bar
    total_row = 14 + len(sorted_stats)
    ws1.row_dimensions[total_row].height = 22
    ws1.merge_cells(f'B{total_row}:H{total_row}')
    _set(ws1, total_row, 2, f'Total: {len(sorted_stats)} estudiantes  ·  Promedio general: {prom_curso}/5',
         bold=True, size=9, color=C['white'], h='center', bg=C['dark'])
    for col in range(1, 9):
        ws1.cell(row=total_row, column=col).fill = _fill(C['dark'])

    # Leyenda
    lg = total_row + 2
    ws1.row_dimensions[lg].height = 14
    _set(ws1, lg, 2, 'Leyenda:', bold=True, size=8, color=C['gray'])
    ws1.merge_cells(f'B{lg}:H{lg}')
    lg += 1
    ws1.row_dimensions[lg].height = 14
    ws1.merge_cells(f'C{lg}:H{lg}')
    ws1.cell(row=lg, column=2).fill = _fill(C['red_l'])
    _set(ws1, lg, 2, '█', size=8, color=C['red'], bg=C['red_l'])
    _set(ws1, lg, 3, 'Estudiante en riesgo: promedio ≤ 2.0 o con alertas activas', size=8, color=C['gray'])
    lg += 1
    ws1.row_dimensions[lg].height = 14
    ws1.merge_cells(f'C{lg}:H{lg}')
    _set(ws1, lg, 2, '█', size=8, color=C['green'], bg=C['green_l'])
    _set(ws1, lg, 3, 'Estudiante destacado: promedio ≥ 4.0 con tendencia positiva', size=8, color=C['gray'])

    # ── Hoja 2: Gráficas ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Gráficas')
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.tabColor = 'FF6B9D'
    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 12

    # Data hidden behind charts
    _set(ws2, 1, 1, 'Estudiante', bold=True, size=9, bg=C['dark'], color=C['white'])
    _set(ws2, 1, 2, 'Promedio', bold=True, size=9, bg=C['dark'], color=C['white'])
    _set(ws2, 1, 3, 'Días Bajos', bold=True, size=9, bg=C['dark'], color=C['white'])

    for i, s in enumerate(sorted_stats):
        r = i + 2
        ws2.cell(row=r, column=1).value = s['nombre'].split()[0]
        ws2.cell(row=r, column=2).value = s['promedio']
        ws2.cell(row=r, column=3).value = s['dias_bajos']

    n = len(sorted_stats)

    # Gráfica de barras — promedio
    bar1 = BarChart()
    bar1.type = 'col'
    bar1.title = 'Promedio Emocional por Estudiante'
    bar1.y_axis.title = 'Puntaje (1–5)'
    bar1.style = 10
    bar1.width = 22
    bar1.height = 14
    bar1.y_axis.scaling.min = 0
    bar1.y_axis.scaling.max = 5
    bar1.y_axis.majorGridlines = None

    d1 = Reference(ws2, min_col=2, min_row=1, max_row=n+1)
    cats = Reference(ws2, min_col=1, min_row=2, max_row=n+1)
    bar1.add_data(d1, titles_from_data=True)
    bar1.set_categories(cats)
    bar1.series[0].graphicalProperties.solidFill = '7C6EFF'
    bar1.series[0].graphicalProperties.line.solidFill = '6B5FEE'
    ws2.add_chart(bar1, 'E2')

    # Gráfica de barras — días bajos
    bar2 = BarChart()
    bar2.type = 'col'
    bar2.title = 'Días con Estado Emocional Bajo (puntaje ≤ 2)'
    bar2.y_axis.title = 'Cantidad de días'
    bar2.style = 10
    bar2.width = 22
    bar2.height = 14

    d2 = Reference(ws2, min_col=3, min_row=1, max_row=n+1)
    bar2.add_data(d2, titles_from_data=True)
    bar2.set_categories(cats)
    bar2.series[0].graphicalProperties.solidFill = 'E0284F'
    bar2.series[0].graphicalProperties.line.solidFill = 'C0002A'
    ws2.add_chart(bar2, 'E22')

    # ── Hoja 3: Detalle por estudiante ───────────────────────────────────────
    ws3 = wb.create_sheet('Detalle Estudiantes')
    ws3.sheet_view.showGridLines = False
    ws3.sheet_properties.tabColor = '00B87A'

    for col, w in zip('ABCDEFGHIJK', [3, 24, 10, 10, 10, 10, 16, 14, 12, 14, 4]):
        ws3.column_dimensions[col].width = w

    # Header
    ws3.merge_cells('B1:K1')
    ws3.row_dimensions[1].height = 30
    _set(ws3, 1, 2, 'Detalle Emocional por Estudiante',
         bold=True, size=13, color=C['white'], h='center', bg=C['dark'])

    h3 = ['', 'Estudiante', 'Promedio', 'Máximo', 'Mínimo', 'Registros',
          'Emoción Dom.', 'Tendencia', 'Racha (días)', 'Días Bajos', 'Alertas']
    ws3.row_dimensions[2].height = 26
    for i, h in enumerate(h3):
        c = ws3.cell(row=2, column=i+1)
        c.value = h
        c.font = _font(bold=True, size=9, color=C['white'])
        c.fill = _fill(C['purple'])
        c.alignment = _align(h='center', wrap=True)
        if i > 0:
            c.border = _border_thin()

    for i, s in enumerate(sorted(stats_lista, key=lambda x: -x['promedio'])):
        row = 3 + i
        ws3.row_dimensions[row].height = 22
        bg = _row_bg(s, i)
        pc = _prom_color(s['promedio'])

        # Acento
        ws3.cell(row=row, column=1).fill = _fill(
            C['red'] if s['alertas_activas'] > 0
            else C['green'] if s['promedio'] >= 4.0
            else bg)

        _set(ws3, row, 2,  s['nombre'],            bold=True, size=9, h='left', bg=bg, border=_border_bottom())

        c = ws3.cell(row=row, column=3)
        c.value = s['promedio']
        c.number_format = '0.00'
        c.font = _font(bold=True, size=11, color=pc)
        c.fill = _fill(bg)
        c.alignment = _align(h='center')
        c.border = _border_bottom()

        _set(ws3, row, 4,  s['puntaje_max'],        size=9, h='center', bg=bg, border=_border_bottom())
        _set(ws3, row, 5,  s['puntaje_min'],        size=9, h='center', bg=bg, border=_border_bottom())
        _set(ws3, row, 6,  s['total_registros'],    size=9, h='center', bg=bg, border=_border_bottom())
        _set(ws3, row, 7,  EMO_LABELS.get(s['emocion_dominante'], '—'), size=9, h='center', bg=bg, border=_border_bottom())
        _set(ws3, row, 8,  _tend_label(s['tendencia']), size=9, h='center', bg=bg, border=_border_bottom())
        _set(ws3, row, 9,  s['racha'],              size=9, h='center', bg=bg, border=_border_bottom())
        _set(ws3, row, 10, s['dias_bajos'],
             bold=s['dias_bajos'] > 3, size=9,
             color=C['red'] if s['dias_bajos'] > 3 else C['text'],
             h='center', bg=bg, border=_border_bottom())

        if s['alertas_activas'] > 0:
            _set(ws3, row, 11, f"⚠ {s['alertas_activas']}", bold=True, size=9,
                 color=C['red'], h='center', bg=C['red_l'], border=_border_bottom())
        else:
            _set(ws3, row, 11, '✓', size=9, color=C['green'], h='center', bg=bg, border=_border_bottom())

    # ── Sub-tabla: distribución de emociones ─
    emo_keys = ['feliz', 'tranquilo', 'estresado', 'triste', 'enojado']
    start = 3 + len(stats_lista) + 3

    ws3.merge_cells(f'B{start}:K{start}')
    ws3.row_dimensions[start].height = 26
    _set(ws3, start, 2, 'Distribucion de Emociones por Estudiante',
         bold=True, size=12, color=C['white'], h='center', bg=C['dark'])

    emo_headers = ['', 'Estudiante'] + [EMO_LABELS[e] for e in emo_keys] + ['Total', '', '']
    ws3.row_dimensions[start+1].height = 24
    for i, h in enumerate(emo_headers[:10]):
        c = ws3.cell(row=start+1, column=i+1)
        c.value = h
        c.font = _font(bold=True, size=9, color=C['white'])
        c.fill = _fill(C['purple'])
        c.alignment = _align(h='center', wrap=True)
        if i > 0 and i < 9:
            c.border = _border_thin()

    for i, s in enumerate(sorted(stats_lista, key=lambda x: -x['promedio'])):
        row = start + 2 + i
        ws3.row_dimensions[row].height = 20
        bg = C['gray_l'] if i % 2 == 0 else C['white']
        total = sum(s['conteo_emociones'].get(e, 0) for e in emo_keys)

        _set(ws3, row, 2, s['nombre'], size=9, h='left', bg=bg, border=_border_bottom())

        for col_i, emo in enumerate(emo_keys):
            cnt = s['conteo_emociones'].get(emo, 0)
            pct = round(cnt / total * 100) if total > 0 else 0
            c = ws3.cell(row=row, column=3+col_i)
            c.value = f'{cnt} ({pct}%)' if cnt > 0 else '—'
            c.font = _font(size=9, color=EMO_FG[emo] if cnt > 0 else C['muted'])
            c.fill = _fill(EMO_BG[emo] if cnt > 0 else bg)
            c.alignment = _align(h='center')
            c.border = _border_bottom()

        _set(ws3, row, 8, total, bold=True, size=9, h='center', bg=bg, border=_border_bottom())

    # ── Hoja 4: Datos crudos ─────────────────────────────────────────────────
    ws4 = wb.create_sheet('Datos Crudos')
    ws4.sheet_properties.tabColor = '6B7280'

    raw_h = ['Estudiante', 'Usuario', 'Total Reg.', 'Promedio', 'Máximo',
             'Mínimo', 'Emoción Dom.', 'Tendencia Num.', 'Racha', 'Días Bajos', 'Alertas']
    for i, h in enumerate(raw_h):
        c = ws4.cell(row=1, column=i+1)
        c.value = h
        c.font = _font(bold=True, size=9, color=C['white'])
        c.fill = _fill(C['gray'])
        ws4.column_dimensions[get_column_letter(i+1)].width = 18

    for i, s in enumerate(stats_lista):
        row = i + 2
        vals = [s['nombre'], s['username'], s['total_registros'], s['promedio'],
                s['puntaje_max'], s['puntaje_min'], s.get('emocion_dominante', ''),
                s['tendencia'], s['racha'], s['dias_bajos'], s['alertas_activas']]
        bg = C['gray_l'] if i % 2 == 0 else C['white']
        for j, v in enumerate(vals):
            c = ws4.cell(row=row, column=j+1)
            c.value = v
            c.font = _font(size=9)
            c.fill = _fill(bg)
            if j == 3:
                c.number_format = '0.00'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ══════════════════════════════════════════════════════════════════════════════
# EXPORTACIÓN PARA ESTUDIANTES — mis propios datos
# ══════════════════════════════════════════════════════════════════════════════

def generar_excel_estudiante(user):
    """
    Genera un Excel con los datos emocionales del propio estudiante.
    Retorna un BytesIO listo para enviar como respuesta HTTP.
    """
    from .models import RegistroEmocional, EntradaDiario, MetaSemanal

    wb = Workbook()

    # ── Hoja 1: Mis registros emocionales ────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Mis Emociones'
    ws1.sheet_view.showGridLines = False
    ws1.sheet_properties.tabColor = 'FF6B9D'

    for col, w in zip('ABCDEF', [3, 14, 18, 10, 16, 40]):
        ws1.column_dimensions[col].width = w

    # Banner
    for merge in ['B1:F1', 'B2:F2', 'B3:F3']:
        ws1.merge_cells(merge)
    for row in [1, 2, 3]:
        ws1.row_dimensions[row].height = 20 if row != 2 else 30
        for col in range(2, 7):
            ws1.cell(row=row, column=col).fill = _fill(C['dark'])

    _set(ws1, 1, 2, 'CONÉCTATE  —  Mi Diario Emocional', size=10, color=C['purple'], h='center', bg=C['dark'])
    _set(ws1, 2, 2, f'{user.get_full_name()}  (@{user.username})',
         bold=True, size=16, color=C['white'], h='center', bg=C['dark'])
    _set(ws1, 3, 2, f'Exportado el {date.today().strftime("%d/%m/%Y")}',
         size=9, color='ADBBC4', h='center', bg=C['dark'])

    ws1.row_dimensions[4].height = 12

    # Stats rápidas
    registros_all = list(RegistroEmocional.objects.filter(estudiante=user).order_by('fecha'))
    n_reg = len(registros_all)
    prom = round(sum(r.puntaje for r in registros_all) / n_reg, 2) if n_reg else 0
    conteo = Counter(r.emocion for r in registros_all)
    emo_dom = conteo.most_common(1)[0][0] if conteo else None

    kpis = [
        ('B', f'{n_reg}', 'Total Registros', C['purple'], C['purple_l']),
        ('C', f'{prom}/5', 'Promedio', _prom_color(prom), C['yellow_l'] if prom >= 3 else C['red_l']),
        ('D', EMO_LABELS.get(emo_dom, '—'), 'Emoción más frecuente', C['pink'], C['pink_l']),
    ]

    for r in [5, 6, 7, 8]:
        ws1.row_dimensions[r].height = 22 if r in [6, 7] else 8

    for col_l, value, label, fg, bg in kpis:
        end_col = chr(ord(col_l) + 1)
        for merge_row in [5, 6, 7, 8]:
            try: ws1.merge_cells(f'{col_l}{merge_row}:{end_col}{merge_row}')
            except: pass
        for rr in [5, 6, 7]:
            for c_off in range(2):
                ws1.cell(row=rr, column=ord(col_l)-64+c_off).fill = _fill(bg)
        _set(ws1, 5, ord(col_l)-63, label, size=8, color=C['gray'], h='center', bg=bg)
        _set(ws1, 6, ord(col_l)-63, value, bold=True, size=16, color=fg, h='center', bg=bg)
        for c_off in range(2):
            ws1.cell(row=8, column=ord(col_l)-64+c_off).fill = _fill(fg)

    ws1.row_dimensions[9].height = 12

    # Tabla de registros
    ws1.row_dimensions[10].height = 24
    reg_headers = ['', 'Fecha', 'Emoción', 'Puntaje', 'Estado', 'Comentario']
    for i, h in enumerate(reg_headers):
        c = ws1.cell(row=10, column=i+1)
        c.value = h
        c.font = _font(bold=True, size=9, color=C['white'])
        c.fill = _fill(C['dark'])
        c.alignment = _align(h='center' if i != 5 else 'left', wrap=True)
        if i > 0:
            c.border = _border_thin()

    puntaje_labels = {5: 'Excelente', 4: 'Bien', 3: 'Regular', 2: 'Difícil', 1: 'Muy difícil'}

    for i, reg in enumerate(sorted(registros_all, key=lambda x: x.fecha, reverse=True)):
        row = 11 + i
        ws1.row_dimensions[row].height = 20
        bg = EMO_BG.get(reg.emocion, C['gray_l']) if i % 2 == 0 else C['white']
        fg = EMO_FG.get(reg.emocion, C['text'])

        ws1.cell(row=row, column=1).fill = _fill(fg)

        _set(ws1, row, 2, reg.fecha.strftime('%d/%m/%Y'), size=9, h='center', bg=bg, border=_border_bottom())
        _set(ws1, row, 3, EMO_LABELS.get(reg.emocion, reg.emocion), size=9, color=fg, h='center', bg=bg, border=_border_bottom())

        c = ws1.cell(row=row, column=4)
        c.value = reg.puntaje
        c.font = _font(bold=True, size=11, color=_prom_color(reg.puntaje))
        c.fill = _fill(bg)
        c.alignment = _align(h='center')
        c.border = _border_bottom()

        _set(ws1, row, 5, puntaje_labels.get(reg.puntaje, ''), size=9, color=_prom_color(reg.puntaje), h='center', bg=bg, border=_border_bottom())
        _set(ws1, row, 6, reg.comentario or '', size=9, h='left', bg=bg, border=_border_bottom(), wrap=True)

    # ── Hoja 2: Mi Diario ────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Mi Diario')
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.tabColor = '7C6EFF'

    for col, w in zip('ABCDE', [3, 15, 30, 50, 4]):
        ws2.column_dimensions[col].width = w

    ws2.merge_cells('B1:D1')
    ws2.row_dimensions[1].height = 28
    _set(ws2, 1, 2, '📓  Mi Diario Personal', bold=True, size=14, color=C['white'], h='center', bg=C['dark'])

    ws2.row_dimensions[2].height = 24
    diary_headers = ['', 'Fecha', 'Título', 'Contenido', '']
    for i, h in enumerate(diary_headers):
        c = ws2.cell(row=2, column=i+1)
        c.value = h
        c.font = _font(bold=True, size=9, color=C['white'])
        c.fill = _fill(C['purple'])
        c.alignment = _align(h='center')
        if 0 < i < 4:
            c.border = _border_thin()

    for i, entry in enumerate(EntradaDiario.objects.filter(estudiante=user).order_by('-created_at')):
        row = 3 + i
        ws2.row_dimensions[row].height = 40
        bg = C['purple_l'] if i % 2 == 0 else C['white']
        ws2.cell(row=row, column=1).fill = _fill(C['purple'])
        _set(ws2, row, 2, entry.created_at.strftime('%d/%m/%Y'), size=9, h='center', bg=bg, border=_border_bottom())
        _set(ws2, row, 3, entry.titulo, bold=True, size=9, h='left', bg=bg, border=_border_bottom())
        c = ws2.cell(row=row, column=4)
        c.value = entry.contenido
        c.font = _font(size=8)
        c.fill = _fill(bg)
        c.alignment = _align(h='left', wrap=True)
        c.border = _border_bottom()

    # ── Hoja 3: Metas ────────────────────────────────────────────────────────
    ws3 = wb.create_sheet('Mis Metas')
    ws3.sheet_view.showGridLines = False
    ws3.sheet_properties.tabColor = '00B87A'

    for col, w in zip('ABCD', [3, 16, 45, 14]):
        ws3.column_dimensions[col].width = w

    ws3.merge_cells('B1:D1')
    ws3.row_dimensions[1].height = 28
    _set(ws3, 1, 2, 'Mis Metas Semanales', bold=True, size=14, color=C['white'], h='center', bg=C['dark'])

    ws3.row_dimensions[2].height = 24
    for i, h in enumerate(['', 'Semana', 'Meta', 'Estado']):
        c = ws3.cell(row=2, column=i+1)
        c.value = h
        c.font = _font(bold=True, size=9, color=C['white'])
        c.fill = _fill(C['green'])
        c.alignment = _align(h='center')
        if i > 0:
            c.border = _border_thin()

    for i, meta in enumerate(MetaSemanal.objects.filter(estudiante=user).order_by('-semana_inicio')):
        row = 3 + i
        ws3.row_dimensions[row].height = 22
        bg = C['green_l'] if meta.cumplida else (C['gray_l'] if i % 2 == 0 else C['white'])
        ws3.cell(row=row, column=1).fill = _fill(C['green'] if meta.cumplida else C['gray'])
        _set(ws3, row, 2, meta.semana_inicio.strftime('%d/%m/%Y'), size=9, h='center', bg=bg, border=_border_bottom())
        _set(ws3, row, 3, meta.texto, size=9, h='left', bg=bg, border=_border_bottom(), wrap=True)
        estado = 'Cumplida' if meta.cumplida else 'Pendiente'
        _set(ws3, row, 4, estado, bold=meta.cumplida, size=9,
             color=C['green'] if meta.cumplida else C['gray'],
             h='center', bg=bg, border=_border_bottom())

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
